from django.http import HttpResponseRedirect
from django.shortcuts import render
from ..forms import UploadFileForm
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import logging
import sys
sys.path.append('C:/projects/mysite/static/graduation_code_review/graduation_code_review_final.py')
logger = logging.getLogger('pybo')

import sys, io
import os

# 액셀, csv를 다루는 모듈 --> pip install openpyxl
import openpyxl, csv

# 파이썬 글자색 적용 모듈 --> pip install termcolor
from termcolor import colored

os.getcwd()


# In[19]:


# 수강한 과목을 확인하는 함수
def get_my_courses(ws):
    print("Getting your courses...", flush=True)

    index = 6
    ret = []

    while True:
        code = ws['B' + str(index)].value
        credit = ws['E' + str(index)].value
        title = ws['D' + str(index)].value

        # 액셀 파일에서 B행 [학사]가 나오는 부분에서
        # 과목코드가 끝남 코드도 그렇게 설계
        if code == '[학사]':
            break;

        if code == None or credit == None or title == None:
            index += 1
            continue;

        # ret라는 배열에 (과목코드. 학점, 과목명) 튜플을 넣음
        ret.append((code, int(credit), title))
        index += 1

    # lambda 함수를 이용해 학점 순으로 정렬
    ret.sort(key=lambda x: x[1], reverse=True)
    print("Done", flush=True)

    return ret



# In[21]:


def classify_my_course(my_course_index, my_courses):
    # my_course = (code, credit, title)
    my_course = my_courses[my_course_index]

    for category in ["core_english1", "core_english2", "core_math1",
                     "core_experiment", "freshman_seminar",
                     "others1_core", "others1_optional", "software", "others3"]:

        # 내 강좌의 과목코드가 (분류) 강좌에 있다면 딕셔너리 value의 학점 더하기 및 튜플 추가
        if my_course[0] in classified_courses[category]:
            my_classified_courses_credit[category] += my_course[1]
            my_classified_courses[category].append(my_course)

            return True

    for category in ["core_writing", "core_math2", "core_science"]:

        # 컴프를 들었을 때 소기코 학점을 0으로
        if my_course[0] == "GS1401":
            classified_courses_credit["software"] = 0
            courses_text["software"] = "Optional"

        if my_course[0] in classified_courses[category]:

            # 기초 과목을 초과해서 들으면 자유 학점으로 처리
            if my_course[1] + my_classified_courses_credit[category] > classified_courses_credit[category]:
                my_classified_courses_credit["others3"] += my_course[1]
                my_classified_courses["others3"].append(my_course)
            # 초과 안하면 기초 필수 과목으로 분류
            else:
                my_classified_courses_credit[category] += my_course[1]
                my_classified_courses[category].append(my_course)

            return True
    # 이 부분을 이해 못함
    # (HUS 과목 이고 (PPE or other humanity)) or (PPE 이고 other humanity)
    if (my_course[0] in classified_courses["HUS"] and (
            my_course[0] in classified_courses["PPE"] or my_course[0] in classified_courses["other_humanity"])) or (
            my_course[0] in classified_courses["PPE"] and my_course[0] in classified_courses["other_humanity"]):
        # 이해 못함
        my_course = list(my_course)
        my_course[2] += " (*)"
        my_course = tuple(my_course)

    # HUS나 PPE를 카테고리에 넣는다
    for category in ["HUS", "PPE"]:

        if my_course[0] in classified_courses[category]:

            if my_course[1] + my_classified_courses_credit[category] > classified_courses_credit[category]:

                if my_course[1] + my_classified_courses_credit["other_humanity"] > classified_courses_credit[
                    "other_humanity"]:
                    my_classified_courses_credit["others2"] = my_classified_courses_credit["others2"] + my_course[1]
                    my_classified_courses["others2"].append(my_course)

                else:
                    my_classified_courses_credit["other_humanity"] += my_course[1]
                    my_classified_courses["other_humanity"].append(my_course)

            else:
                my_classified_courses_credit[category] += my_course[1]
                my_classified_courses[category].append(my_course)

            return True

    if my_course[0] in classified_courses["other_humanity"]:

        if my_course[1] + my_classified_courses_credit["other_humanity"] > classified_courses_credit["other_humanity"]:
            my_classified_courses_credit["others2"] += my_course[1]
            my_classified_courses["others2"].append(my_course)

        else:
            my_classified_courses_credit["other_humanity"] += my_course[1]
            my_classified_courses["other_humanity"].append(my_course)

        return True

    # Classify my major courses
    for classified_category, category in zip(major[my_major],
                                             ["major_core", "major_elective"]):

        if my_course[0] in classified_courses[classified_category]:
            my_classified_courses_credit[category] += my_course[1]
            my_classified_courses[category].append(my_course)

            return True

    # Classify other major courses
    for classified_category in [category for sublist in major[:my_major] + major[my_major + 1:] for category in
                                sublist]:

        if my_course[0] in classified_courses[classified_category]:
            my_classified_courses_credit["others3"] += my_course[1]
            my_classified_courses["others3"].append(my_course)

            return True

    # Classify research courses
    for code in classified_courses["research"]:

        if my_course[0][2:] == code:
            my_classified_courses_credit["research"] += my_course[1]
            my_classified_courses["research"].append(my_course)

            return True

    for category in ["music", "exercise", "colloquium"]:

        if my_course[0] in classified_courses[category]:
            my_classified_courses_credit[category] += 1
            my_classified_courses[category].append(my_course)

            return True

    return False


def sum_credits():
    ret = 0  # result

    # category라는 변수에 강좌 분류를 넣는다
    for category in ["core_english1", "core_english2", "core_writing",
                     "HUS", "PPE", "other_humanity", "software",
                     "core_math1", "core_math2",
                     "core_science", "core_experiment",
                     "freshman_seminar", "research",
                     "others1_core", "others2"]:
        # 내가 들은 category 강좌를 넣는다 min을 사용하는 이유는 충족해야하는 학점을 넘으면 자유 학점으로 채우기 위함.
        ret += min(my_classified_courses_credit[category],
                   classified_courses_credit[category])

    # 전공 필수+ 선택, 전공 최대 학점 중 최솟값을 출력 (전공보다 더들었을 경우 학점 계산 X를 위함)
    ret += min(my_classified_courses_credit["major_core"] + my_classified_courses_credit["major_elective"],
               classified_courses_credit["major"])
    # 위와 비슷
    ret += (my_classified_courses_credit["others1_optional"] + my_classified_courses_credit["others3"])

    # 총 들은 학점을 출력
    return ret


def print_course(course):
    print('{:<7} {:<7d} {:}'.format(course[0], course[1], course[2]))
    # course = (code, credit, title) ,과목코드 학점 과목명 순ㅇ로 출력


# 0학점 과목 분류 (콜로퀴움, 예체능)
def print_courses_by_class(mclass, length=70):
    # 처음에 과목 분류를 출력 ex) Core English1
    print(colored('\n' + courses_name[mclass].center(length) + '\n',
                  attrs=['bold']))

    print('-' * 75)
    print('{:<7} {:<7} {:}'.format("course", "credit", "title"))

    # 내가 들은 과목들을 출력 ex) GS1601 2 영어 1: 신입생 영어
    for course in my_classified_courses[mclass]:
        print_course(course)

    print('-' * 75)

    # 필수 과목이면 Mandatory 만약 필수과목을 안들었다면 빨간색 굵은 글씨로 출력
    if courses_text[mclass] == "Mandatory" and classified_courses_credit[mclass] > my_classified_courses_credit[mclass]:
        print(colored('{:7} {:<7} {:}'.format(" ",
                                              str(my_classified_courses_credit[mclass]) + '/' + \
                                              str(classified_courses_credit[mclass]),
                                              courses_text[mclass]), "red", attrs=['bold']))

    else:
        print('{:7} {:<7} {:}'.format(" ",
                                      str(my_classified_courses_credit[mclass]) + '/' + \
                                      str(classified_courses_credit[mclass]),
                                      courses_text[mclass]))

    print('-' * 75 + '\n')


# 유학점 분류
def print_courses_by_subclass(subclass):
    print('- ' + courses_name[subclass] + '\n' + '-' * 75)
    print('{:<7} {:<7} {:}'.format("course", "credit", "title"))

    for course in my_classified_courses[subclass]:
        print_course(course)

    print('-' * 75)

    if subclass == "core_experiment":
        output_string = '{:7} {:<7} {:}'.format(" ",
                                                str(my_classified_courses_credit[subclass]) + '/' + \
                                                str(classified_courses_credit[subclass]),
                                                courses_text[subclass])

        if classified_courses_credit[subclass] < 2:
            print(colored(output_string, "red", attrs=['bold']) + '\n' + '-' * 75 + '\n')
            return
        else:
            print(output_string + '\n' + '-' * 75 + '\n')
            return

    output_string = '{:7} {:<7} {:}'.format(" ",
                                            str(my_classified_courses_credit[subclass]) + '/' + \
                                            str(classified_courses_credit[subclass]),
                                            courses_text[subclass])
    if courses_text[subclass] == "Mandatory" and classified_courses_credit[subclass] > my_classified_courses_credit[
        subclass]:
        print(colored(output_string, "red", attrs=['bold']))
    else:
        print(output_string)

    print('-' * 75 + '\n')


# 전공 과목 출력
def print_major_courses():
    print(colored('\n' + "전공학점".center(70) + '\n',
                  attrs=['bold']))

    print("- Major Core")
    print('-' * 75)
    print('{:<7} {:<7} {:}'.format("course", "credit", "title"))
    for course in my_classified_courses["major_core"]:
        print_course(course)

    print('-' * 75)
    print('{:7} {:<7}'.format(" ",
                              str(my_classified_courses_credit["major_core"])))
    print('-' * 75)

    print("\n- Major elective")
    print('-' * 75)
    print('{:<7} {:<7} {:}'.format("course", "credit", "title"))
    for course in my_classified_courses["major_elective"]:
        print_course(course)

    print('-' * 75)
    print('{:7} {:<7}'.format(" ",
                              str(my_classified_courses_credit["major_elective"])))
    print('-' * 75)

    major_credit = my_classified_courses_credit["major_elective"] + my_classified_courses_credit["major_core"]
    output_string = '{:7} {:<7} {:}'.format(" ",
                                            str(major_credit) + '/' + \
                                            str(classified_courses_credit["major"]),
                                            "Mandatory over 36")

    if major_credit < 36:
        print(colored(output_string, "red", attrs=['bold']))
    else:
        print(output_string)

    print('-' * 75 + '\n')


# 강의 딕셔너리
"""
classified_course = [core_english1, core_english2, core_writing,
        HUS, PPE, other_humanity, software,
        core_math1, core_math2, core_science, core_experiment,
        freshman_seminar,
        physics_core, physics_elective,
        chemical_core, chemical_elective,
        biology_core, biology_elective,
        eecs_core, eecs_elective,
        mechanics_core, mechanics_elective,
        environment_core, environment_elective,
        research,
        others1_core, others1_optional, others3,
        music, exercise, colloquium]
"""
classified_courses = {
    'core_english1': ['GS1601', 'GS1603'],
    'core_english2': ['GS2652'],
    'core_writing': ['GS1511', 'GS1512', 'GS1513', 'GS1531', 'GS1532', 'GS1533'],
    'HUS': ['GS2501', 'GS2503', 'GS2505', 'GS2506', 'GS2507', 'GS2509', 'GS2510', 'GS2511', 'GS2512', 'GS2521',
            'GS2522', 'GS2523', 'GS2524', 'GS2525', 'GS2526', 'GS2544', 'GS2601', 'GS2602', 'GS2603', 'GS2604',
            'GS2611', 'GS2612', 'GS2613', 'GS2614', 'GS2615', 'GS2616', 'GS2618', 'GS2621', 'GS2622', 'GS2623',
            'GS2625', 'GS2626', 'GS2627', 'GS2628', 'GS2629', 'GS2630', 'GS2814', 'GS3501', 'GS3502', 'GS3504',
            'GS3601', 'GS3602', 'GS3603', 'GS3604', 'GS3621', 'GS3622', 'GS3623', 'GS3624', 'GS3625', 'GS3626',
            'GS3662', 'GS3801', 'GS3802', 'GS3803', 'GS3901'],
    'PPE': ['GS2620', 'GS2661', 'GS2701', 'GS2702', 'GS2703', 'GS2704', 'GS2705', 'GS2706', 'GS2707', 'GS2708',
            'GS2709', 'GS2724', 'GS2725', 'GS2726', 'GS2727', 'GS2728', 'GS2729', 'GS2730', 'GS2731', 'GS2732',
            'GS2733', 'GS2734', 'GS2735', 'GS2736', 'GS2742', 'GS2743', 'GS2747', 'GS2748', 'GS2750', 'GS2751',
            'GS2752', 'GS2761', 'GS2762', 'GS2763', 'GS2764', 'GS2765', 'GS2766', 'GS2781', 'GS2782', 'GS2783',
            'GS2784', 'GS2785', 'GS2786', 'GS2787', 'GS2788', 'GS2803', 'GS2812', 'GS2831', 'GS2832', 'GS2833',
            'GS2834', 'GS2835', 'GS3631', 'GS3632', 'GS3633', 'GS3661', 'GS3663', 'GS3721', 'GS3751', 'GS3752',
            'GS3753', 'GS3762', 'GS3763', 'GS3764', 'GS3861', 'GS4741', 'GS4761', 'GS4762'],
    'other_humanity': ['GS2541', 'GS2542', 'GS2543', 'GS2544', 'GS2791', 'GS2792', 'GS2793', 'GS2804', 'GS2808',
                       'GS2810', 'GS2815', 'GS2816', 'GS2817', 'GS2818', 'GS2819', 'GS2821', 'GS2822', 'GS2911',
                       'GS2912', 'GS2913', 'GS2931', 'GS2932', 'GS2933', 'GS3566'],
    'software': ['GS1490'],
    'core_math1': ['GS1001', 'GS1011'],
    'core_math2': ['GS2001', 'GS2002', 'GS2004', 'GS2013'],
    'core_science': ['GS1101', 'GS1103', 'GS1201', 'GS1203', 'GS1301', 'GS1302', 'GS1303', 'GS1401'],
    'core_experiment': ['GS1111', 'GS1211', 'GS1311'],
    'freshman_seminar': ['GS1901', 'GS9301'],
    'physics_core': ['PS2101', 'PS2102', 'PS2103', 'PS3101', 'PS3103', 'PS3104', 'PS3105', 'PS3106', 'PS3107'],
    'physics_elective': ['PS2201', 'PS2202', 'PS3202', 'PS3203', 'PS3205', 'PS3206', 'PS4202', 'PS4203', 'PS4204',
                         'PS4205', 'PS4206', 'PS4207', 'PS4208', 'PS4209', 'PS4210', 'PS4211', 'PS4212', 'PS4213',
                         'PS4214', 'PS4215', 'PS4216'],
    'chemical_core': ['CH2101', 'CH2102', 'CH2103', 'CH2104', 'CH2105', 'CH3102', 'CH3103', 'CH3104', 'CH3105',
                      'CH3106', 'CH3107'],
    'chemical_elective': ['CH2106', 'CH2201', 'CH3202', 'CH3204', 'CH3205', 'CH3207', 'CH4205', 'CH4211', 'CH4212',
                          'CH4213', 'CH4215', 'CH4216', 'CH4218', 'CH4219', 'CH4220', 'CH4221', 'CH4222', 'CH4223'],
    'biology_core': ['BS2101', 'BS2102', 'BS3101', 'BS3105', 'BS3111', 'BS3112', 'BS3113'],
    'biology_elective': ['BS2201', 'BS3201', 'BS3202', 'BS3204', 'BS3205', 'BS4201', 'BS4202', 'BS4204', 'BS4205',
                         'BS4206', 'BS4207', 'BS4211', 'BS4212', 'BS4213', 'BS4214', 'BS4215', 'BS4216', 'BS4217',
                         'BS4218'],
    'eecs_core': ['EC3101', 'EC3102'],
    'eecs_elective': ['EC2105', 'EC2201', 'EC2202', 'EC2203', 'EC2204', 'EC2205', 'EC2206', 'EC3102', 'EC3202',
                      'EC3204', 'EC3206', 'EC3207', 'EC3208', 'EC3212', 'EC3213', 'EC3214', 'EC3215', 'EC3216',
                      'EC3217', 'EC3218', 'EC4202', 'EC4203', 'EC4204', 'EC4205', 'EC4206', 'EC4207', 'EC4208',
                      'EC4209', 'EC4210', 'EC4211', 'EC4212', 'EC4213', 'EC4214', 'EC4215', 'EC4216', 'EC4217',
                      'EC4218', 'EC4219', 'EC4301', 'EC4302'],
    'mechanics_core': ['MC2100', 'MC2101', 'MC2102', 'MC2103', 'MC3103', 'MC3105', 'MC3212', 'MC4101'],
    'mechanics_elective': ['MC3201', 'MC3202', 'MC3203', 'MC3204', 'MC3205', 'MC3206', 'MC3207', 'MC3208', 'MC3209',
                           'MC3210', 'MC3211', 'MC4202', 'MC4204', 'MC4205', 'MC4206', 'MC4208', 'MC4209', 'MC4210',
                           'MC4211', 'MC4212', 'MC4213', 'MC4214', 'MC4215', 'MC4216', 'MC4217', 'MC4218', 'MC4219',
                           'MC4221'],
    'material_core': ['MA2101', 'MA2102', 'MA2103', 'MA2104', 'MA3101', 'MA3102', 'MA3104', 'MA3105'],
    'material_elective': ['MA2201', 'MA2202', 'MA3201', 'MA3202', 'MA3203', 'MA3204', 'MA3205', 'MA3207', 'MA3208',
                          'MA3209', 'MA3210', 'MA3211', 'MA4201', 'MA4202', 'MA4203', 'MA4204', 'MA4205', 'MA4206',
                          'MA4207', 'MA4208', 'MA4209', 'MA4210', 'MA4211', 'MA4212', 'MA4213', 'MA4214', 'MA4215',
                          'MA4216', 'MA4217', 'MA4218', 'MA4219', 'MA4220', 'MA4221'],
    'environment_core': ['EV3101', 'EV3104', 'EV3105', 'EV3106', 'EV3111', 'EV4105', 'EV4106', 'EV4107'],
    'environment_elective': ['EV2208', 'EV2209', 'EV2210', 'EV2211', 'EV3205', 'EV3208', 'EV3213', 'EV3214', 'EV3215',
                             'EV3216', 'EV3217', 'EV3218', 'EV3219', 'EV3220', 'EV4201', 'EV4202', 'EV4203', 'EV4204',
                             'EV4205', 'EV4206', 'EV4209', 'EV4210', 'EV4211', 'EV4212', 'EV4213', 'EV4214', 'EV4215',
                             'EV4216', 'EV4217', 'EV4218', 'EV4221', 'EV4222', 'EV4223', 'EV4224', 'EV4225'],
    'research': ['9102', '9103', '9104'],
    'others1_core': ['UC0901'],
    'others1_optional': ['UC0201', 'UC0202', 'UC0203', 'UC0301'],
    'others3': ['CT2501', 'CT2502', 'CT2503', 'CT2504', 'CT2505', 'CT2506', 'CT4101', 'CT41__', 'CT4201', 'CT4301',
                'CT4302', 'CT4501', 'CT4502', 'CT4503', 'CT4504', 'CT4506', 'CT45__', 'ET2101', 'ET4102', 'ET4201',
                'ET4302', 'ET4303', 'ET4304', 'ET4305', 'ET4306', 'ET4501', 'GS1102', 'GS1104', 'GS1112', 'GS1202',
                'GS1204', 'GS1212', 'GS1321', 'GS1402', 'GS1431', 'GS1451', 'GS1471', 'GS1491', 'GS1605', 'GS1606',
                'GS2006', 'GS2007', 'GS2102', 'GS2103', 'GS2104', 'GS2201', 'GS2202', 'GS2204', 'GS2206', 'GS2302',
                'GS2303', 'GS2304', 'GS2311', 'GS2401', 'GS2402', 'GS2403', 'GS2406', 'GS2407', 'GS2408', 'GS2434',
                'GS2435', 'GS2451', 'GS2471', 'GS2472', 'GS2473', 'GS2651', 'GS2653', 'GS2654', 'GS2655', 'GS2806',
                'GS2809', 'GS2811', 'GS3001', 'GS3012', 'GS3015', 'GS3301', 'GS3311', 'GS3651', 'GS4002', 'GS4003',
                'GS4004', 'GS4005', 'GS4006', 'GS4007', 'GS4008', 'GS4009', 'GS4010', 'GS4015', 'GS4016', 'GS4017',
                'GS4018', 'GS4019', 'GS4301', 'IR2201', 'IR2202', 'IR3201', 'IR3202', 'IR3203', 'IR3204', 'IR4201',
                'IR4202', 'IR4203', 'IR4204', 'IR4205', 'IR4206', 'IR4207', 'IR4208', 'IR4209', 'MD2101', 'MD4101',
                'MD4102', 'MD4301', 'MD4302', 'MD4303', 'MD4501', 'MD4502', 'MD4601', 'MM2001', 'MM2002', 'MM2004',
                'MM2006', 'MM2007', 'MM2011', 'MM3001', 'MM3012', 'MM3015', 'MM4002', 'MM4003', 'MM4004', 'MM4005',
                'MM4006', 'MM4007', 'MM4008', 'MM4009', 'MM4010', 'MM4015', 'MM4016', 'MM4017', 'MM4018', 'MM4019',
                'others3', 'others3', 'others3', 'others3', 'others3'],
    'exercise': ['GS0101', 'GS0102', 'GS0103', 'GS0104', 'GS0105', 'GS0106', 'GS0107', 'GS0108', 'GS0109', 'GS0110',
                 'GS0111', 'GS0112', 'GS0113', 'GS0114'],
    'music': ['GS0201', 'GS0202', 'GS0203', 'GS0204', 'GS0205', 'GS0206', 'GS0207', 'GS0208', 'GS0209', 'GS0210',
              'GS0211', 'GS0212'],
    'colloquium': ['GS9331', 'UC9331']
}

# 졸업을 위해 들어야하는 학점을 넣는 딕셔너리
classified_courses_credit = {
    "core_english1": 2, "core_english2": 2, "core_writing": 3,
    "HUS": 6, "PPE": 6, "other_humanity": 12, "software": 2,
    "core_math1": 3, "core_math2": 3,
    "core_science": 9, "core_experiment": 3,
    "freshman_seminar": 1,
    "major": 42,
    "research": 6,
    "others1_core": 1, "others1_optional": '-', "others2": 12, "others3": '-',
    "music": 4, "exercise": 4, "colloquium": 2,
    "nonclassified_courses": '-'
}
# 필수, 선택 연구 학점 분류 하는 딕셔너리
courses_text = {
    **dict.fromkeys(
        ["core_english1", "core_english2", "core_writing",
         "HUS", "PPE", "other_humanity",
         "core_math1", "core_math2", "core_science",
         "freshman_seminar",
         "research", "others1_core",
         "music", "exercise", "colloquium"],
        "Mandatory"),
    **dict.fromkeys(
        ["others1_optional", "others2", "others3"], "Optional"),
    "software": "Mandatory",
    "core_experiment": "Mandatory over 2",
    "nonclassified_courses": ""
}
# 강좌 이름 딕셔너리
courses_name = {
    "core_english1": "Core English 1",
    "core_english2": "Core English 2",
    "core_writing": "Core Writing",
    "HUS": "HUS", "PPE": "PPE",
    "other_humanity": "Other Humanity",
    "software": "소프트웨어",
    "core_math1": "Core Mathematics 1",
    "core_math2": "Core Mathematics 2",
    "core_science": "Core Science",
    "core_experiment": "Core Experiment",
    "freshman_seminar": "신입생 세미나",
    "research": "학사논문연구",
    "others1_core": "Mandatory",
    "others1_optional": "Optional",
    "others2": "자유선택 - 인문사회",
    "others3": "자유선택 - 언어선택/소프트웨어, \
기초과학선택, 기초전공, 타 전공",
    "music": "예능실기", "exercise": "체육실기",
    "colloquium": "GIST대학 콜로퀴움",
    "nonclassified_courses": "Nonclassified Courses\
 (학사편람에 기재되지 않은 과목)"
}
major = [["physics_core", "physics_elective"],
         ["chemical_core", "chemical_elective"],
         ["biology_core", "biology_elective"],
         ["eecs_core", "eecs_elective"],
         ["mechanics_core", "mechanics_elective"],
         ["material_core", "material_elective"],
         ["environment_core", "environment_elective"]]

"""
my_classified_course = [core_english1, core_english2, core_writing,
        HUS, PPE, other_humanity,
        core_math1, core_math2, core_science, core_experiment,
        freshman_seminar,
        major_core, major_elective,
        research,
        others1, other2, others3,
        music, exercise, colloquium, nonclassified_courses]
"""
"""
0: physics, 1: chemical, 2: biology, 3: eecs,
4: mechanics, 5: materials, 6: environment
"""

# 이 부분을 체크
my_major = 3

# 내가 들은 강좌들을 담는 딕셔너리
my_classified_courses = {category: [] for category in [
    "core_english1", "core_english2", "core_writing",
    "HUS", "PPE", "other_humanity", "software",
    "core_math1", "core_math2", "core_science", "core_experiment",
    "freshman_seminar",
    "major_core", "major_elective",
    "research",
    "others1_core", "others1_optional", "others2", "others3",
    "music", "exercise", "colloquium", "nonclassified_courses"]}

# 내가 들은 학점들을 담는 딕셔너리
my_classified_courses_credit = {
    **dict.fromkeys(
        ["core_english1", "core_english2", "core_writing",
         "HUS", "PPE", "other_humanity", "software",
         "core_math1", "core_math2", "core_science", "core_experiment",
         "freshman_seminar",
         "major_core", "major_elective",
         "research",
         "others1_core", "others1_optional", "others2", "others3",
         "music", "exercise", "colloquium", "nonclassified_courses"], 0)}

# my_courses = list of (code, credit, title)
# my_courses라는 변수에 내가 들은 강의를 넣는다


def upload_file(request):
    if request.method == 'POST' and request.FILES['document']:
        loaded_file = request.FILES['document']
        print(loaded_file.name)
        print(loaded_file.size)
        print(loaded_file.file)
        fs = FileSystemStorage()
        fs.save(loaded_file.name, loaded_file.file, max_length=None)
        print('-' * 70)
        print("Enter the number corresponding with your major".center(70))
        print("0: Physics, 1: Chemical, 2: Biology, 3: EECS,".center(70))
        print("4: Mechanics, 5: Materials, 6: environment".center(70))
        print('-' * 70)
        ws = openpyxl.load_workbook('C:/projects/mysite/media/Completed course grade.xlsx').active
        my_courses = get_my_courses(ws)
        my_major = 3

        for my_course_index in range(len(my_courses)):

            if not classify_my_course(my_course_index, my_courses):
                my_classified_courses_credit["nonclassified_courses"] += my_courses[my_course_index][1]
                my_classified_courses["nonclassified_courses"].append(
                    my_courses[my_course_index])

        print(colored("\n" + "언어의 기초".center(70) + "\n",
                      attrs=['bold']))
        for subclass in ["core_english1", "core_english2", "core_writing"]:
            print_courses_by_subclass(subclass)

        print(colored("\n" + "인문사회".center(70) + "\n",
                      attrs=['bold']))
        for subclass in ["HUS", "PPE", "other_humanity"]:
            print_courses_by_subclass(subclass)

        print_courses_by_class("software")

        print(colored("\n" + "기초과학".center(70) + "\n",
                      attrs=['bold']))
        for subclass in ["core_math1", "core_math2", "core_science",
                         "core_experiment"]:
            print_courses_by_subclass(subclass)

        print_courses_by_class("freshman_seminar")
        print_major_courses()

        print_courses_by_class("research", 70)

        print(colored("\n" + "자유선택 - 대학 공통과목".center(70) + "\n",
                      attrs=['bold']))
        for subclass in ["others1_core", "others1_optional"]:
            print_courses_by_subclass(subclass)

        print_courses_by_class("others2", 69)
        print_courses_by_class("others3", 50)

        print('\n' + colored(('=' * 31).center(75), "red", attrs=['bold']))
        print(colored("||      Total Credits      ||".center(75), "red",
                      attrs=['bold']))
        print(
            colored(("{:}{:^25}{:}".format("||", str(sum_credits()) + "/130", "||")).center(75), "red", attrs=['bold']))
        print(colored(('=' * 31).center(75), "red", attrs=['bold']))
        print()

        print_courses_by_class("nonclassified_courses", length=63)
        for mclass in ["music", "exercise", "colloquium"]:
            print_courses_by_class(mclass)

        print(my_classified_courses)
        file = 'C:/projects/mysite/media/Completed course grade.xlsx'
        if os.path.isfile(file):
            os.remove(file)
    return render(request, 'pybo/upload.html')